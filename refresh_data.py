#!/usr/bin/env python3
"""Refresh data.json from a Financial Flash .xlsx export.

Usage:
    python refresh_data.py "Financial Flash.xlsx"

Pulls these rows from the Summary tab, columns E:EU (147 weekly columns):
    Row 242 - Pieces Produced (volume)
    Row 251 - Production Hours Worked
    Row 252 - Fulfillment Hours Worked
    Row 253 - Total Hours Worked
    Row 256 - Pieces Produced / Production Hr. Worked
    Row 257 - Net Pieces Shipped / FF Hours Worked
    Row 258 - Pieces Produced / Total Hours Worked

LY-aligned values are computed by shifting 52 weeks (same fiscal week prior year).
Wage rate of $25.53/hr is applied for cost-savings calc.
"""
import sys, json, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError:
    sys.exit("Install openpyxl first:  pip install openpyxl")

WAGE = 25.53
SHEET = "Summary"
COL_FROM = "E"
COL_TO = "EU"

ROW_DEFS = {
    "volume_242":        (242, "Pieces Produced (volume)"),
    "prod_hours_251":    (251, "Production Hours Worked"),
    "ff_hours_252":      (252, "Fulfillment Hours Worked"),
    "total_hours_253":   (253, "Total Hours Worked"),
    "eff_prod_256":      (256, "Pieces Produced / Production Hr. Worked"),
    "eff_ff_257":        (257, "Net Pieces Shipped / FF Hours Worked"),
    "eff_total_258":     (258, "Pieces Produced / Total Hours Worked"),
}
HEADER_ROWS = {"fy": 4, "month": 5, "week_start": 6, "week_end": 7, "retail_calendar": 9}


def iso(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return v


def num(v):
    if v is None:
        return None
    if isinstance(v, str):
        try:
            return float(v.replace(",", ""))
        except ValueError:
            return None
    return float(v)


def main(xlsx_path: str, out_path: str = "data.json"):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Sheet {SHEET!r} not found. Tabs: {wb.sheetnames}")
    ws = wb[SHEET]

    e = column_index_from_string(COL_FROM)
    eu = column_index_from_string(COL_TO)
    n = eu - e + 1

    needed = set(HEADER_ROWS.values()) | {r for r, _ in ROW_DEFS.values()}
    rows = {}
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=max(needed), max_col=eu, values_only=True), start=1):
        if i in needed:
            rows[i] = list(r)[e - 1: eu]

    weeks = []
    for c in range(n):
        w = {"idx": c + 1}
        for k, r in HEADER_ROWS.items():
            v = rows[r][c]
            w[k] = iso(v) if k in ("week_start", "week_end") else v
        for key, (r, _) in ROW_DEFS.items():
            w[key] = num(rows[r][c])
        weeks.append(w)

    # LY alignment + savings
    for c, w in enumerate(weeks):
        ly_idx = c - 52
        if ly_idx >= 0:
            ly = weeks[ly_idx]
            w["ly_eff_total_258"] = ly["eff_total_258"]
            w["ly_volume_242"] = ly["volume_242"]
            w["ly_total_hours_253"] = ly["total_hours_253"]
            if w["volume_242"] is not None and ly["eff_total_258"] not in (None, 0):
                w["hours_at_ly_eff"] = w["volume_242"] / ly["eff_total_258"]
            else:
                w["hours_at_ly_eff"] = None
            if w["hours_at_ly_eff"] is not None and w["total_hours_253"] is not None:
                w["hours_saved_vs_ly"] = w["hours_at_ly_eff"] - w["total_hours_253"]
                w["wage_savings_vs_ly"] = w["hours_saved_vs_ly"] * WAGE
            else:
                w["hours_saved_vs_ly"] = None
                w["wage_savings_vs_ly"] = None
        else:
            for k in ("ly_eff_total_258", "ly_volume_242", "ly_total_hours_253",
                     "hours_at_ly_eff", "hours_saved_vs_ly", "wage_savings_vs_ly"):
                w[k] = None

    snapshot = {
        "source": "Financial Flash (Summary tab) - rows 242, 251-253, 256-258 across cols E:EU",
        "snapshot_date": datetime.date.today().isoformat(),
        "row_definitions": {str(r): label for _, (r, label) in ROW_DEFS.items()},
        "wage_per_hour": WAGE,
        "weeks": weeks,
    }

    Path(out_path).write_text(json.dumps(snapshot, indent=2, default=str))
    print(f"Wrote {out_path} ({len(weeks)} weeks)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Usage: python refresh_data.py <path-to-Financial-Flash.xlsx> [data.json]")
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "data.json")
